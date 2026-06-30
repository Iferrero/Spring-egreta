import openpyxl

wb = openpyxl.load_workbook("src/main/resources/TABLON-Balanç de recursos - detallat (RC0025R).xlsx", read_only=True)
print("=== SHEETS ===")
print(wb.sheetnames)

sheet = wb.active
print(f"=== ACTIVE SHEET TITLE: {sheet.title} ===")

# Print first 20 rows of first sheet to see headers
print("\n=== FIRST 20 ROWS ===")
row_count = 0
for row in sheet.iter_rows(values_only=True):
    print(row[:15])
    row_count += 1
    if row_count >= 20:
        break

wb.close()
