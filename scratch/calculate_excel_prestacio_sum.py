import openpyxl

wb = openpyxl.load_workbook("src/main/resources/TABLON-Balanç de recursos - detallat (RC0025R).xlsx", read_only=True)
sheet = wb.active

total_2024 = 0.0
count_2024 = 0

for row in sheet.iter_rows(min_row=2, values_only=True):
    # Year is in column 2 (1-based index 1) or let's inspect columns first
    # In AwardService.java, columns are parsed. Let's see:
    # row[0] is ID, row[4] is year? Let's print headers
    if row[0] is None:
        continue
    # Let's inspect first row of data
    # print(row)
    # break

# Let's read columns like AwardService.java:
# AwardService.java:
# Row 0 is header.
# Col 6 is "Any" or "exercici"?
# Let's print headers
headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
print("Excel Headers:", headers)

# Find year column
year_idx = -1
for i, h in enumerate(headers):
    if h and "exercici" in h.lower():
        year_idx = i
        break
    if h and "any" in h.lower():
        year_idx = i
        break

# Find value column
val_idx = -1
for i, h in enumerate(headers):
    if h and "import" in h.lower():
        val_idx = i
        break

print(f"Year Index: {year_idx}, Value Index: {val_idx}")

# Calculate
for row in sheet.iter_rows(min_row=2, values_only=True):
    if len(row) > max(year_idx, val_idx):
        yr = row[year_idx]
        val = row[val_idx]
        if yr == 2024 and val is not None:
            total_2024 += float(val)
            count_2024 += 1

print(f"2024 Total Prestacions: {total_2024 / 1e6:.2f} M€ (Count: {count_2024})")
wb.close()
