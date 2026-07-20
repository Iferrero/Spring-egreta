import openpyxl

wb = openpyxl.load_workbook(r"c:\Users\1000788\OneDrive - UAB\Documentos\GitHub\Spring-egreta\src\main\resources\Indicadors resum_2025.xlsx", data_only=True)
print("Sheets in Indicadors resum_2025.xlsx:")
print(wb.sheetnames)

for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    print(f"\n--- Sheet: {sheetname} (Dimensions: {sheet.dimensions}) ---")
    # Print the first 10 rows and first 15 columns
    max_r = min(sheet.max_row, 30)
    max_c = min(sheet.max_column, 25)
    for r in range(1, max_r + 1):
        row_vals = []
        for c in range(1, max_c + 1):
            val = sheet.cell(row=r, column=c).value
            row_vals.append(val)
        # Check if row is not completely empty
        if any(v is not None for v in row_vals):
            print(f"Row {r:02d}: {row_vals}")
