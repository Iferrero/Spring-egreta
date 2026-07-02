import openpyxl
import os

file_path = r"c:\Users\1000788\OneDrive - UAB\Documentos\GitHub\Spring-egreta\src\main\resources\Trams de recerca vius de la UAB (RC0019).xlsx"
if not os.path.exists(file_path):
    print("File not found")
else:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    print("Active sheet:", sheet.title)
    
    for r in range(1, sheet.max_row + 1):
        row_vals = [sheet.cell(r, c).value for c in range(1, 10)]
        if any(row_vals):
            print(f"Row {r}: {row_vals}")
